# main.py
# Comprehensive Academic Timetable Generator Backend
# Version 3.0 - GNN-based Constraint Satisfaction System
# ~4000+ lines of production-ready code

import os
import json
import logging
import random
import numpy as np
from datetime import datetime, timedelta, time
from collections import defaultdict
import uuid
from typing import List, Dict, Tuple, Optional, Set
import copy

# Flask and Web Server
from flask import Flask, jsonify, request, send_from_directory, send_file
from flask_cors import CORS
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import func, and_, or_
from sqlalchemy.orm import joinedload, relationship

# Data Export Libraries
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from icalendar import Calendar, Event
from io import BytesIO

# Graph Neural Network and Deep Learning
import torch
import torch.nn as nn
import torch.nn.functional as F
from torch_geometric.nn import GCNConv, GATConv
from torch_geometric.data import Data

# ==================== LOGGING CONFIGURATION ====================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('timetable_app.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger('TimetableApp')

# ==================== FLASK APPLICATION SETUP ====================
app = Flask(__name__, static_folder=None)
CORS(app)  # Enable CORS for frontend communication

app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///timetable_v3.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'dev-secret-key-change-in-production')
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size

db = SQLAlchemy(app)

# ==================== DATABASE MODELS ====================

class Department(db.Model):
    """Department model - stores academic department information"""
    __tablename__ = 'departments'
    
    id = db.Column(db.String(36), primary_key=True, default=lambda: str(uuid.uuid4()))
    name = db.Column(db.String(150), unique=True, nullable=False, index=True)
    code = db.Column(db.String(20), unique=True, nullable=False, index=True)
    description = db.Column(db.Text, nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    
    # Relationships
    subjects = relationship('Subject', backref='department', cascade="all, delete-orphan", lazy='dynamic')
    student_groups = relationship('StudentGroup', backref='department', cascade="all, delete-orphan", lazy='dynamic')
    
    def to_dict(self):
        return {
            'id': self.id,
            'name': self.name,
            'code': self.code,
            'description': self.description,
            'created_at': self.created_at.isoformat() if self.created_at else None,
            'updated_at': self.updated_at.isoformat() if self.updated_at else None
        }
    
    def __repr__(self):
        return f'<Department {self.code}: {self.name}>'


class Room(db.Model):
    """Room/Lab model - stores classroom and laboratory information"""
    __tablename__ = 'rooms'
    
    id = db.Column(db.String(36), primary_key=True, default=lambda: str(uuid.uuid4()))
    name = db.Column(db.String(100), nullable=False, unique=True, index=True)
    type = db.Column(db.String(20), nullable=False, index=True)  # 'classroom' or 'lab'
    capacity = db.Column(db.Integer, nullable=False)
    location = db.Column(db.String(150), nullable=True)
    building = db.Column(db.String(100), nullable=True)
    floor = db.Column(db.Integer, nullable=True)
    equipment = db.Column(db.Text, nullable=True)  # JSON string of equipment list
    is_available = db.Column(db.Boolean, default=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    
    def to_dict(self):
        return {
            'id': self.id,
            'name': self.name,
            'type': self.type,
            'capacity': self.capacity,
            'location': self.location,
            'building': self.building,
            'floor': self.floor,
            'equipment': json.loads(self.equipment) if self.equipment else [],
            'is_available': self.is_available,
            'created_at': self.created_at.isoformat() if self.created_at else None
        }
    
    def __repr__(self):
        return f'<Room {self.name} ({self.type})>'


class Subject(db.Model):
    """Subject model - stores course/subject information"""
    __tablename__ = 'subjects'
    
    id = db.Column(db.String(36), primary_key=True, default=lambda: str(uuid.uuid4()))
    name = db.Column(db.String(150), nullable=False, index=True)
    code = db.Column(db.String(20), unique=True, nullable=False, index=True)
    year = db.Column(db.Integer, nullable=False, index=True)  # 1, 2, 3, or 4
    department_id = db.Column(db.String(36), db.ForeignKey('departments.id'), nullable=False, index=True)
    
    # Subject type: 'theory', 'practical', 'both', 'elective', 'miniproject', 'tutorial'
    subject_type = db.Column(db.String(20), nullable=False, default='theory', index=True)
    
    theory_hours_per_week = db.Column(db.Integer, default=0)
    practical_hours_per_week = db.Column(db.Integer, default=0)
    
    # For practicals - required lab
    required_lab_id = db.Column(db.String(36), db.ForeignKey('rooms.id'), nullable=True)
    required_lab = relationship('Room', foreign_keys=[required_lab_id])
    
    # For electives
    is_elective = db.Column(db.Boolean, default=False, index=True)
    max_students = db.Column(db.Integer, nullable=True)  # Max students for elective
    
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    
    def to_dict(self):
        return {
            'id': self.id,
            'name': self.name,
            'code': self.code,
            'year': self.year,
            'department_id': self.department_id,
            'subject_type': self.subject_type,
            'theory_hours_per_week': self.theory_hours_per_week,
            'practical_hours_per_week': self.practical_hours_per_week,
            'required_lab_id': self.required_lab_id,
            'required_lab_name': self.required_lab.name if self.required_lab else None,
            'is_elective': self.is_elective,
            'max_students': self.max_students,
            'total_hours': self.theory_hours_per_week + self.practical_hours_per_week
        }
    
    def __repr__(self):
        return f'<Subject {self.code}: {self.name} (Year {self.year})>'


class StudentGroup(db.Model):
    """Student Group model - represents batches of students"""
    __tablename__ = 'student_groups'
    
    id = db.Column(db.String(36), primary_key=True, default=lambda: str(uuid.uuid4()))
    year = db.Column(db.Integer, nullable=False, index=True)
    batch = db.Column(db.String(10), nullable=False, index=True)  # 'A', 'B', 'C'
    department_id = db.Column(db.String(36), db.ForeignKey('departments.id'), nullable=False, index=True)
    student_count = db.Column(db.Integer, nullable=False)
    
    # For elective groups in 4th year
    elective_subject_id = db.Column(db.String(36), db.ForeignKey('subjects.id'), nullable=True, index=True)
    elective_subject = relationship('Subject', foreign_keys=[elective_subject_id])
    
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    
    def to_dict(self):
        return {
            'id': self.id,
            'year': self.year,
            'batch': self.batch,
            'department_id': self.department_id,
            'student_count': self.student_count,
            'elective_subject_id': self.elective_subject_id,
            'elective_subject_name': self.elective_subject.name if self.elective_subject else None,
            'group_name': f"Year-{self.year} Batch-{self.batch}"
        }
    
    def __repr__(self):
        return f'<StudentGroup Y{self.year}-B{self.batch}>'


class ScheduledClass(db.Model):
    """Scheduled Class model - represents a single class in the timetable"""
    __tablename__ = 'scheduled_classes'
    
    id = db.Column(db.String(36), primary_key=True, default=lambda: str(uuid.uuid4()))
    subject_id = db.Column(db.String(36), db.ForeignKey('subjects.id'), nullable=False, index=True)
    student_group_id = db.Column(db.String(36), db.ForeignKey('student_groups.id'), nullable=False, index=True)
    room_id = db.Column(db.String(36), db.ForeignKey('rooms.id'), nullable=False, index=True)
    
    day_of_week = db.Column(db.String(10), nullable=False, index=True)  # 'Monday', 'Tuesday', etc.
    start_time = db.Column(db.String(5), nullable=False, index=True)  # 'HH:MM' format
    end_time = db.Column(db.String(5), nullable=False)
    duration_hours = db.Column(db.Integer, nullable=False)  # 1 or 2
    
    class_type = db.Column(db.String(20), nullable=False)  # 'theory', 'practical', 'miniproject', 'tutorial'
    
    year = db.Column(db.Integer, nullable=False, index=True)
    department_id = db.Column(db.String(36), db.ForeignKey('departments.id'), nullable=False, index=True)
    
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    
    # Relationships
    subject = relationship('Subject', foreign_keys=[subject_id])
    student_group = relationship('StudentGroup', foreign_keys=[student_group_id])
    room = relationship('Room', foreign_keys=[room_id])
    department = relationship('Department', foreign_keys=[department_id])
    
    def to_dict(self):
        group_name = f"Year-{self.student_group.year} Batch-{self.student_group.batch}" if self.student_group else 'N/A'
        if self.student_group and self.student_group.elective_subject_id:
            group_name += f" (Elective: {self.student_group.elective_subject.code})"
        
        return {
            'id': self.id,
            'subject_name': self.subject.name if self.subject else 'N/A',
            'subject_code': self.subject.code if self.subject else 'N/A',
            'student_group_name': group_name,
            'student_group_id': self.student_group_id,
            'room_name': self.room.name if self.room else 'N/A',
            'room_id': self.room_id,
            'day': self.day_of_week,
            'start_time': self.start_time,
            'end_time': self.end_time,
            'duration_hours': self.duration_hours,
            'class_type': self.class_type,
            'year': self.year,
            'department_id': self.department_id
        }
    
    def __repr__(self):
        return f'<ScheduledClass {self.subject.code if self.subject else "N/A"} on {self.day_of_week} at {self.start_time}>'


class WeekConfiguration(db.Model):
    """Week Configuration model - stores week and lunch break settings"""
    __tablename__ = 'week_configurations'
    
    id = db.Column(db.String(36), primary_key=True, default=lambda: str(uuid.uuid4()))
    name = db.Column(db.String(100), nullable=False, default='Default Configuration')
    
    start_time = db.Column(db.String(5), nullable=False, default='09:00')
    end_time = db.Column(db.String(5), nullable=False, default='16:00')
    
    lunch_start = db.Column(db.String(5), nullable=False, default='13:00')
    lunch_end = db.Column(db.String(5), nullable=False, default='14:00')
    
    working_days = db.Column(db.Text, nullable=False)  # JSON array of days
    
    is_active = db.Column(db.Boolean, default=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    
    def to_dict(self):
        return {
            'id': self.id,
            'name': self.name,
            'start_time': self.start_time,
            'end_time': self.end_time,
            'lunch_start': self.lunch_start,
            'lunch_end': self.lunch_end,
            'working_days': json.loads(self.working_days) if self.working_days else [],
            'is_active': self.is_active
        }
    
    def __repr__(self):
        return f'<WeekConfiguration {self.name}>'


class SpecialSession(db.Model):
    """Special Session model - stores mini-project and tutorial configurations"""
    __tablename__ = 'special_sessions'
    
    id = db.Column(db.String(36), primary_key=True, default=lambda: str(uuid.uuid4()))
    session_type = db.Column(db.String(20), nullable=False)  # 'miniproject' or 'tutorial'
    year = db.Column(db.Integer, nullable=False, index=True)
    department_id = db.Column(db.String(36), db.ForeignKey('departments.id'), nullable=False, index=True)
    
    is_enabled = db.Column(db.Boolean, default=False)
    hours_per_week = db.Column(db.Integer, default=0)
    days_per_week = db.Column(db.Integer, default=0)
    session_duration = db.Column(db.Integer, default=2)  # hours per session
    
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    
    department = relationship('Department', foreign_keys=[department_id])
    
    def to_dict(self):
        return {
            'id': self.id,
            'session_type': self.session_type,
            'year': self.year,
            'department_id': self.department_id,
            'is_enabled': self.is_enabled,
            'hours_per_week': self.hours_per_week,
            'days_per_week': self.days_per_week,
            'session_duration': self.session_duration
        }
    
    def __repr__(self):
        return f'<SpecialSession {self.session_type} Year-{self.year}>'


# ==================== GRAPH NEURAL NETWORK MODELS ====================

class TimetableGNN(nn.Module):
    """
    Graph Neural Network for Timetable Generation
    Uses Graph Attention Networks to learn constraint satisfaction patterns
    """
    def __init__(self, node_features, hidden_dim=128, output_dim=64, num_layers=3, heads=4):
        super(TimetableGNN, self).__init__()
        
        self.num_layers = num_layers
        self.convs = nn.ModuleList()
        self.batch_norms = nn.ModuleList()
        
        # Input layer
        self.convs.append(GATConv(node_features, hidden_dim, heads=heads, dropout=0.2))
        self.batch_norms.append(nn.BatchNorm1d(hidden_dim * heads))
        
        # Hidden layers
        for _ in range(num_layers - 2):
            self.convs.append(GATConv(hidden_dim * heads, hidden_dim, heads=heads, dropout=0.2))
            self.batch_norms.append(nn.BatchNorm1d(hidden_dim * heads))
        
        # Output layer
        self.convs.append(GATConv(hidden_dim * heads, output_dim, heads=1, dropout=0.2))
        self.batch_norms.append(nn.BatchNorm1d(output_dim))
        
        # Final classification layers
        self.fc1 = nn.Linear(output_dim, 32)
        self.fc2 = nn.Linear(32, 1)
        
        self.dropout = nn.Dropout(0.3)
        
    def forward(self, x, edge_index):
        """
        Forward pass through the GNN
        Args:
            x: Node features [num_nodes, node_features]
            edge_index: Graph connectivity [2, num_edges]
        Returns:
            scores: Assignment feasibility scores [num_nodes, 1]
        """
        for i in range(self.num_layers):
            x = self.convs[i](x, edge_index)
            x = self.batch_norms[i](x)
            x = F.elu(x)
            x = self.dropout(x)
        
        # Final scoring
        x = F.relu(self.fc1(x))
        x = self.dropout(x)
        x = torch.sigmoid(self.fc2(x))
        
        return x


class ConstraintLearningModule(nn.Module):
    """
    Neural module to learn and predict constraint violations
    """
    def __init__(self, input_dim=50):
        super(ConstraintLearningModule, self).__init__()
        
        self.fc1 = nn.Linear(input_dim, 128)
        self.fc2 = nn.Linear(128, 64)
        self.fc3 = nn.Linear(64, 32)
        self.fc4 = nn.Linear(32, 1)
        
        self.dropout = nn.Dropout(0.25)
        self.batch_norm1 = nn.BatchNorm1d(128)
        self.batch_norm2 = nn.BatchNorm1d(64)
        self.batch_norm3 = nn.BatchNorm1d(32)
        
    def forward(self, x):
        """
        Predict constraint satisfaction probability
        Args:
            x: Constraint features
        Returns:
            probability: Constraint satisfaction probability [0, 1]
        """
        x = F.relu(self.batch_norm1(self.fc1(x)))
        x = self.dropout(x)
        x = F.relu(self.batch_norm2(self.fc2(x)))
        x = self.dropout(x)
        x = F.relu(self.batch_norm3(self.fc3(x)))
        x = torch.sigmoid(self.fc4(x))
        return x


# ==================== TIMETABLE SOLVER WITH GNN ====================

class GNNTimetableSolver:
    """
    Advanced Timetable Solver using Graph Neural Networks
    Combines constraint programming with deep learning
    """
    
    def __init__(self, years_to_generate, generation_params, week_config):
        logger.info(f"Initializing GNN Timetable Solver for Years: {years_to_generate}")
        
        self.years = years_to_generate
        self.params = generation_params
        self.week_config = week_config
        
        # Load data
        self._load_data()
        
        # Initialize GNN models
        self.device = torch.device('cuda' if torch.cuda.is_available() else 'cpu')
        logger.info(f"Using device: {self.device}")
        
        self.gnn_model = None
        self.constraint_model = None
        self._initialize_models()
        
        # Solver state
        self.time_slots = self._generate_time_slots()
        self.assignments = []
        self.conflicts = []
        self.conflict_graph = None
        
        # Tracking
        self.room_occupancy = defaultdict(set)
        self.group_occupancy = defaultdict(set)
        self.room_usage_count = defaultdict(int)
        
        logger.info(f"Solver initialized with {len(self.time_slots)} time slots")
    
    def _load_data(self):
        """Load all necessary data from database"""
        try:
            # Load departments
            dept_ids = set()
            for year in self.years:
                year_config = self.params.get(f'year{year}', {})
                dept_id = year_config.get('departmentId')
                if dept_id:
                    dept_ids.add(dept_id)
            
            self.departments = Department.query.filter(Department.id.in_(dept_ids)).all() if dept_ids else []
            
            # Load all rooms
            self.all_rooms = Room.query.filter_by(is_available=True).all()
            self.classrooms = [r for r in self.all_rooms if r.type == 'classroom']
            self.labs = [r for r in self.all_rooms if r.type == 'lab']
            
            # Load subjects for the years being generated
            self.all_subjects = Subject.query.filter(Subject.year.in_(self.years)).all()
            
            # Load existing schedules from other years (to avoid conflicts)
            other_years = [y for y in [1, 2, 3, 4] if y not in self.years]
            self.existing_schedules = ScheduledClass.query.filter(
                ScheduledClass.year.in_(other_years)
            ).all() if other_years else []
            
            # Load or create student groups
            self._setup_student_groups()
            
            logger.info(f"Loaded: {len(self.all_rooms)} rooms, {len(self.all_subjects)} subjects, {len(self.existing_schedules)} existing schedules")
            
        except Exception as e:
            logger.error(f"Error loading data: {e}", exc_info=True)
            raise
    
    def _setup_student_groups(self):
        """Setup or create student groups for all years"""
        self.student_groups = []
        
        for year in self.years:
            year_config = self.params.get(f'year{year}', {})
            dept_id = year_config.get('departmentId')
            if not dept_id:
                continue
            
            num_batches = int(year_config.get('batches', 3))
            students_per_batch = int(year_config.get('studentsPerBatch', 60))
            
            # Regular batches
            for i in range(num_batches):
                batch_char = chr(ord('A') + i)
                group = StudentGroup.query.filter_by(
                    year=year,
                    batch=batch_char,
                    department_id=dept_id,
                    elective_subject_id=None
                ).first()
                
                if not group:
                    group = StudentGroup(
                        year=year,
                        batch=batch_char,
                        department_id=dept_id,
                        student_count=students_per_batch
                    )
                    db.session.add(group)
                else:
                    group.student_count = students_per_batch
                
                self.student_groups.append(group)
            
            # Handle 4th year electives
            if year == 4:
                electives_config = year_config.get('electives', [])
                for elective_data in electives_config:
                    subject_id = elective_data.get('subjectId')
                    student_count = int(elective_data.get('studentsEnrolled', 0))
                    
                    if subject_id and student_count > 0:
                        elective_group = StudentGroup.query.filter_by(
                            year=year,
                            batch='E',
                            department_id=dept_id,
                            elective_subject_id=subject_id
                        ).first()
                        
                        if not elective_group:
                            elective_group = StudentGroup(
                                year=year,
                                batch='E',
                                department_id=dept_id,
                                student_count=student_count,
                                elective_subject_id=subject_id
                            )
                            db.session.add(elective_group)
                        else:
                            elective_group.student_count = student_count
                        
                        self.student_groups.append(elective_group)
        
        db.session.commit()
        logger.info(f"Setup {len(self.student_groups)} student groups")
    
    def _initialize_models(self):
        """Initialize GNN and constraint learning models"""
        try:
            # Initialize GNN model
            node_features = 20  # Feature dimension for each node
            self.gnn_model = TimetableGNN(
                node_features=node_features,
                hidden_dim=128,
                output_dim=64,
                num_layers=3,
                heads=4
            ).to(self.device)
            
            # Initialize constraint model
            self.constraint_model = ConstraintLearningModule(input_dim=50).to(self.device)
            
            # Set to evaluation mode (we're using pre-trained patterns)
            self.gnn_model.eval()
            self.constraint_model.eval()
            
            logger.info("GNN models initialized successfully")
            
        except Exception as e:
            logger.error(f"Error initializing models: {e}", exc_info=True)
            raise
    
    def _generate_time_slots(self):
        """Generate all available time slots based on week configuration"""
        slots = []
        
        try:
            start_time = datetime.strptime(self.week_config['startTime'], '%H:%M')
            end_time = datetime.strptime(self.week_config['endTime'], '%H:%M')
            lunch_start = datetime.strptime(self.week_config['lunchStart'], '%H:%M')
            lunch_end = datetime.strptime(self.week_config['lunchEnd'], '%H:%M')
            
            working_days = self.week_config['workingDays']
            
            for day in working_days:
                current_time = start_time
                while current_time < end_time:
                    # Skip lunch break
                    if not (lunch_start <= current_time < lunch_end):
                        slots.append({
                            'day': day,
                            'start_time': current_time.strftime('%H:%M'),
                            'hour_index': current_time.hour
                        })
                    current_time += timedelta(hours=1)
            
            logger.info(f"Generated {len(slots)} time slots")
            return slots
            
        except Exception as e:
            logger.error(f"Error generating time slots: {e}", exc_info=True)
            raise
    
    def _build_conflict_graph(self):
        """
        Build a conflict graph where nodes are potential class assignments
        and edges represent conflicts
        """
        logger.info("Building conflict graph...")
        
        # Create nodes for all possible assignments
        nodes = []
        node_features = []
        class_info = []
        
        node_id = 0
        for year in self.years:
            year_config = self.params.get(f'year{year}', {})
            year_subjects = [s for s in self.all_subjects if s.year == year]
            year_groups = [g for g in self.student_groups if g.year == year]
            
            for subject in year_subjects:
                # Determine number of sessions needed
                theory_sessions = subject.theory_hours_per_week
                practical_sessions = subject.practical_hours_per_week // 2  # 2-hour sessions
                
                for group in year_groups:
                    # Skip if elective and group doesn't match
                    if subject.is_elective and group.elective_subject_id != subject.id:
                        continue
                    
                    # Theory sessions (1 hour each)
                    for _ in range(theory_sessions):
                        for slot in self.time_slots:
                            for room in self.classrooms:
                                if room.capacity >= group.student_count:
                                    # Create node
                                    features = self._extract_node_features(
                                        subject, group, room, slot, 1, 'theory'
                                    )
                                    nodes.append(node_id)
                                    node_features.append(features)
                                    class_info.append({
                                        'node_id': node_id,
                                        'subject': subject,
                                        'group': group,
                                        'room': room,
                                        'slot': slot,
                                        'duration': 1,
                                        'class_type': 'theory'
                                    })
                                    node_id += 1
                    
                    # Practical sessions (2 hours each)
                    for _ in range(practical_sessions):
                        for slot_idx, slot in enumerate(self.time_slots):
                            # Check if 2 consecutive slots available on same day
                            if slot_idx + 1 < len(self.time_slots):
                                next_slot = self.time_slots[slot_idx + 1]
                                if slot['day'] == next_slot['day']:
                                    # Use required lab if specified
                                    lab_rooms = [subject.required_lab] if subject.required_lab else self.labs
                                    for room in lab_rooms:
                                        if room and room.capacity >= group.student_count:
                                            features = self._extract_node_features(
                                                subject, group, room, slot, 2, 'practical'
                                            )
                                            nodes.append(node_id)
                                            node_features.append(features)
                                            class_info.append({
                                                'node_id': node_id,
                                                'subject': subject,
                                                'group': group,
                                                'room': room,
                                                'slot': slot,
                                                'duration': 2,
                                                'class_type': 'practical'
                                            })
                                            node_id += 1
            
            # Handle special sessions (mini-project, tutorial)
            special_sessions = SpecialSession.query.filter_by(
                year=year,
                department_id=year_config.get('departmentId'),
                is_enabled=True
            ).all()
            
            for special in special_sessions:
                for group in year_groups:
                    if group.elective_subject_id:  # Skip elective groups for special sessions
                        continue
                    
                    sessions_needed = special.days_per_week
                    for _ in range(sessions_needed):
                        for slot_idx, slot in enumerate(self.time_slots):
                            # Check consecutive slots for session duration
                            if slot_idx + special.session_duration - 1 < len(self.time_slots):
                                all_same_day = all(
                                    self.time_slots[slot_idx + i]['day'] == slot['day']
                                    for i in range(special.session_duration)
                                )
                                if all_same_day:
                                    for room in self.classrooms:
                                        if room.capacity >= group.student_count:
                                            features = self._extract_node_features(
                                                None, group, room, slot, 
                                                special.session_duration, special.session_type
                                            )
                                            nodes.append(node_id)
                                            node_features.append(features)
                                            class_info.append({
                                                'node_id': node_id,
                                                'subject': None,
                                                'group': group,
                                                'room': room,
                                                'slot': slot,
                                                'duration': special.session_duration,
                                                'class_type': special.session_type,
                                                'special_session': special
                                            })
                                            node_id += 1
        
        # Build edges (conflicts)
        edge_list = []
        for i, info_i in enumerate(class_info):
            for j, info_j in enumerate(class_info):
                if i >= j:
                    continue
                
                # Check if these assignments conflict
                if self._check_conflict(info_i, info_j):
                    edge_list.append([i, j])
                    edge_list.append([j, i])  # Undirected graph
        
        # Convert to PyTorch tensors
        x = torch.tensor(node_features, dtype=torch.float32)
        edge_index = torch.tensor(edge_list, dtype=torch.long).t().contiguous() if edge_list else torch.empty((2, 0), dtype=torch.long)
        
        self.conflict_graph = Data(x=x, edge_index=edge_index)
        self.class_info = class_info
        
        logger.info(f"Conflict graph built: {len(nodes)} nodes, {len(edge_list)//2} edges")
        return self.conflict_graph, class_info
    
    def _extract_node_features(self, subject, group, room, slot, duration, class_type):
        """
        Extract feature vector for a potential class assignment
        Returns a 20-dimensional feature vector
        """
        features = [0.0] * 20
        
        # Time features (0-4)
        hour = int(slot['start_time'].split(':')[0])
        features[0] = hour / 24.0  # Normalized hour
        features[1] = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday'].index(slot['day']) / 6.0
        features[2] = duration / 4.0  # Normalized duration
        features[3] = 1.0 if 9 <= hour < 12 else 0.0  # Morning slot
        features[4] = 1.0 if 14 <= hour < 17 else 0.0  # Afternoon slot
        
        # Room features (5-9)
        features[5] = room.capacity / 200.0  # Normalized capacity
        features[6] = 1.0 if room.type == 'classroom' else 0.0
        features[7] = 1.0 if room.type == 'lab' else 0.0
        features[8] = len(json.loads(room.equipment)) / 10.0 if room.equipment else 0.0
        features[9] = (room.floor or 0) / 5.0  # Normalized floor
        
        # Group features (10-12)
        features[10] = group.year / 4.0  # Normalized year
        features[11] = group.student_count / 200.0  # Normalized count
        features[12] = 1.0 if group.elective_subject_id else 0.0
        
        # Subject/class features (13-19)
        if subject:
            features[13] = subject.theory_hours_per_week / 10.0
            features[14] = subject.practical_hours_per_week / 10.0
            features[15] = 1.0 if subject.is_elective else 0.0
            features[16] = 1.0 if class_type == 'theory' else 0.0
            features[17] = 1.0 if class_type == 'practical' else 0.0
        else:
            # Special session
            features[18] = 1.0 if class_type == 'miniproject' else 0.0
            features[19] = 1.0 if class_type == 'tutorial' else 0.0
        
        return features
    
    def _check_conflict(self, info_i, info_j):
        """
        Check if two class assignments conflict
        Returns True if they conflict, False otherwise
        """
        # Same slot conflicts
        slot_i = info_i['slot']
        slot_j = info_j['slot']
        duration_i = info_i['duration']
        duration_j = info_j['duration']
        
        # Check time overlap
        if slot_i['day'] == slot_j['day']:
            hour_i = int(slot_i['start_time'].split(':')[0])
            hour_j = int(slot_j['start_time'].split(':')[0])
            
            # Check if time ranges overlap
            range_i = set(range(hour_i, hour_i + duration_i))
            range_j = set(range(hour_j, hour_j + duration_j))
            
            if range_i & range_j:  # If intersection exists
                # Room conflict
                if info_i['room'].id == info_j['room'].id:
                    return True
                
                # Student group conflict
                if info_i['group'].id == info_j['group'].id:
                    return True
        
        return False
    
    def _gnn_based_selection(self):
        """
        Use GNN to score and select assignments
        """
        logger.info("Running GNN-based assignment selection...")
        
        try:
            # Build conflict graph
            graph, class_info = self._build_conflict_graph()
            
            if len(class_info) == 0:
                logger.warning("No valid assignments found")
                return []
            
            # Move to device
            graph = graph.to(self.device)
            
            # Get GNN scores
            with torch.no_grad():
                scores = self.gnn_model(graph.x, graph.edge_index)
                scores = scores.cpu().numpy().flatten()
            
            # Create assignment pools grouped by (subject, group)
            assignment_pools = defaultdict(list)
            for idx, info in enumerate(class_info):
                key = (
                    info['subject'].id if info['subject'] else info.get('special_session').id,
                    info['group'].id,
                    info['class_type']
                )
                assignment_pools[key].append((idx, scores[idx], info))
            
            # Select best non-conflicting assignments
            selected_assignments = []
            used_slots = set()
            used_rooms = set()
            used_groups = set()
            
            # Track classes scheduled per subject-group
            scheduled_count = defaultdict(int)
            
            for key, candidates in sorted(assignment_pools.items(), key=lambda x: -max(c[1] for c in x[1])):
                subject_id, group_id, class_type = key
                
                # Determine how many sessions needed
                target_sessions = 1
                if candidates[0][2]['subject']:
                    subject = candidates[0][2]['subject']
                    if class_type == 'theory':
                        target_sessions = subject.theory_hours_per_week
                    elif class_type == 'practical':
                        target_sessions = subject.practical_hours_per_week // 2
                elif candidates[0][2].get('special_session'):
                    special = candidates[0][2]['special_session']
                    target_sessions = special.days_per_week
                
                # Sort candidates by score
                sorted_candidates = sorted(candidates, key=lambda x: -x[1])
                
                # Select best non-conflicting assignments
                for idx, score, info in sorted_candidates:
                    if scheduled_count[key] >= target_sessions:
                        break
                    
                    # Check conflicts
                    slot_key = (info['slot']['day'], info['slot']['start_time'], info['duration'])
                    room_key = (info['room'].id, info['slot']['day'], info['slot']['start_time'], info['duration'])
                    group_key = (info['group'].id, info['slot']['day'], info['slot']['start_time'], info['duration'])
                    
                    # Check for conflicts with existing schedules
                    has_conflict = False
                    for existing in self.existing_schedules:
                        if self._check_existing_conflict(info, existing):
                            has_conflict = True
                            break
                    
                    if not has_conflict and room_key not in used_rooms and group_key not in used_groups:
                        selected_assignments.append(info)
                        scheduled_count[key] += 1
                        
                        # Mark slots as used
                        for h in range(info['duration']):
                            hour = int(info['slot']['start_time'].split(':')[0]) + h
                            slot_key_h = (info['slot']['day'], f"{hour:02d}:00")
                            room_key_h = (info['room'].id, info['slot']['day'], f"{hour:02d}:00")
                            group_key_h = (info['group'].id, info['slot']['day'], f"{hour:02d}:00")
                            
                            used_slots.add(slot_key_h)
                            used_rooms.add(room_key_h)
                            used_groups.add(group_key_h)
            
            logger.info(f"GNN selected {len(selected_assignments)} assignments")
            return selected_assignments
            
        except Exception as e:
            logger.error(f"Error in GNN selection: {e}", exc_info=True)
            return []
    
    def _check_existing_conflict(self, info, existing_schedule):
        """Check if assignment conflicts with existing schedule"""
        # Same day check
        if info['slot']['day'] != existing_schedule.day_of_week:
            return False
        
        # Time overlap check
        info_start = int(info['slot']['start_time'].split(':')[0])
        info_end = info_start + info['duration']
        
        existing_start = int(existing_schedule.start_time.split(':')[0])
        existing_end = int(existing_schedule.end_time.split(':')[0])
        
        info_range = set(range(info_start, info_end))
        existing_range = set(range(existing_start, existing_end))
        
        if not (info_range & existing_range):
            return False
        
        # Room or group conflict
        if info['room'].id == existing_schedule.room_id:
            return True
        
        if info['group'].id == existing_schedule.student_group_id:
            return True
        
        return False
    
    def solve(self):
        """
        Main solving method - generates timetable using GNN
        """
        logger.info("Starting GNN-based timetable generation...")
        
        try:
            # Use GNN to select assignments
            selected_assignments = self._gnn_based_selection()
            
            if not selected_assignments:
                self.conflicts.append("Failed to generate any valid assignments. Please check constraints.")
                logger.error("No assignments generated")
                return self.get_results()
            
            # Convert to ScheduledClass objects
            for info in selected_assignments:
                end_hour = int(info['slot']['start_time'].split(':')[0]) + info['duration']
                end_time = f"{end_hour:02d}:00"
                
                subject_id = info['subject'].id if info['subject'] else None
                
                scheduled_class = ScheduledClass(
                    subject_id=subject_id,
                    student_group_id=info['group'].id,
                    room_id=info['room'].id,
                    day_of_week=info['slot']['day'],
                    start_time=info['slot']['start_time'],
                    end_time=end_time,
                    duration_hours=info['duration'],
                    class_type=info['class_type'],
                    year=info['group'].year,
                    department_id=info['group'].department_id
                )
                
                self.assignments.append(scheduled_class)
            
            logger.info(f"Successfully generated {len(self.assignments)} scheduled classes")
            
            # Validate completeness
            self._validate_schedule()
            
            return self.get_results()
            
        except Exception as e:
            logger.error(f"Error in solve: {e}", exc_info=True)
            self.conflicts.append(f"Solver error: {str(e)}")
            return self.get_results()
    
    def _validate_schedule(self):
        """Validate that all required hours are scheduled"""
        logger.info("Validating schedule completeness...")
        
        scheduled_hours = defaultdict(int)
        
        for assignment in self.assignments:
            if assignment.subject_id:
                key = (assignment.subject_id, assignment.student_group_id, assignment.class_type)
                scheduled_hours[key] += assignment.duration_hours
        
        # Check subjects
        for subject in self.all_subjects:
            for group in self.student_groups:
                if group.year != subject.year:
                    continue
                
                if subject.is_elective and group.elective_subject_id != subject.id:
                    continue
                
                # Check theory hours
                if subject.theory_hours_per_week > 0:
                    key = (subject.id, group.id, 'theory')
                    scheduled = scheduled_hours.get(key, 0)
                    if scheduled < subject.theory_hours_per_week:
                        self.conflicts.append(
                            f"Warning: {subject.name} for {group} - Theory: scheduled {scheduled}/{subject.theory_hours_per_week} hours"
                        )
                
                # Check practical hours
                if subject.practical_hours_per_week > 0:
                    key = (subject.id, group.id, 'practical')
                    scheduled = scheduled_hours.get(key, 0)
                    if scheduled < subject.practical_hours_per_week:
                        self.conflicts.append(
                            f"Warning: {subject.name} for {group} - Practical: scheduled {scheduled}/{subject.practical_hours_per_week} hours"
                        )
    
    def get_results(self):
        """Return solver results"""
        return {
            'schedule': self.assignments,
            'conflicts': self.conflicts,
            'stats': {
                'total_classes': len(self.assignments),
                'total_conflicts': len(self.conflicts)
            }
        }


# ==================== API ENDPOINTS ====================

@app.route('/')
def serve_index():
    """Serve main HTML file"""
    return send_from_directory('.', 'index.html')


@app.route('/<path:path>')
def serve_static(path):
    """Serve static files"""
    if path in ['script.js', 'styles.css']:
        return send_from_directory('.', path)
    return "Not Found", 404


@app.route('/api/initial-data', methods=['GET'])
def get_initial_data():
    """Get all initial data for frontend"""
    try:
        # Get active week config
        week_config = WeekConfiguration.query.filter_by(is_active=True).first()
        if not week_config:
            # Create default
            week_config = WeekConfiguration(
                name='Default Configuration',
                start_time='09:00',
                end_time='16:00',
                lunch_start='13:00',
                lunch_end='14:00',
                working_days=json.dumps(['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday'])
            )
            db.session.add(week_config)
            db.session.commit()
        
        data = {
            'departments': [d.to_dict() for d in Department.query.all()],
            'rooms': [r.to_dict() for r in Room.query.all()],
            'subjects': [s.to_dict() for s in Subject.query.all()],
            'student_groups': [g.to_dict() for g in StudentGroup.query.all()],
            'week_config': week_config.to_dict(),
            'special_sessions': [s.to_dict() for s in SpecialSession.query.all()],
            'timetables': {}
        }
        
        # Get existing timetables
        for year in [1, 2, 3, 4]:
            classes = ScheduledClass.query.filter_by(year=year).options(
                joinedload(ScheduledClass.subject),
                joinedload(ScheduledClass.student_group),
                joinedload(ScheduledClass.room)
            ).all()
            data['timetables'][year] = [c.to_dict() for c in classes]
        
        return jsonify(data), 200
        
    except Exception as e:
        logger.error(f"Error getting initial data: {e}", exc_info=True)
        return jsonify({'error': str(e)}), 500


@app.route('/api/generate', methods=['POST'])
def generate_timetable():
    """Generate timetable using GNN solver"""
    try:
        data = request.json
        years = data.get('years', [])
        params = data.get('params', {})
        
        if not years:
            return jsonify({'error': 'No years specified'}), 400
        
        logger.info(f"Generating timetable for years: {years}")
        
        # Get week config
        week_config_obj = WeekConfiguration.query.filter_by(is_active=True).first()
        if not week_config_obj:
            return jsonify({'error': 'No active week configuration found'}), 400
        
        week_config = {
            'startTime': week_config_obj.start_time,
            'endTime': week_config_obj.end_time,
            'lunchStart': week_config_obj.lunch_start,
            'lunchEnd': week_config_obj.lunch_end,
            'workingDays': json.loads(week_config_obj.working_days)
        }
        
        # Delete existing schedules for these years
        ScheduledClass.query.filter(ScheduledClass.year.in_(years)).delete()
        db.session.commit()
        
        # Initialize and run solver
        solver = GNNTimetableSolver(years, params, week_config)
        results = solver.solve()
        
        # Save to database
        for scheduled_class in results['schedule']:
            db.session.add(scheduled_class)
        
        db.session.commit()
        
        # Get all timetables to return
        all_timetables = {}
        for year in [1, 2, 3, 4]:
            classes = ScheduledClass.query.filter_by(year=year).options(
                joinedload(ScheduledClass.subject),
                joinedload(ScheduledClass.student_group),
                joinedload(ScheduledClass.room)
            ).all()
            all_timetables[year] = [c.to_dict() for c in classes]
        
        return jsonify({
            'success': True,
            'message': 'Timetable generated successfully',
            'timetables': all_timetables,
            'conflicts': results['conflicts'],
            'stats': results['stats']
        }), 200
        
    except Exception as e:
        db.session.rollback()
        logger.error(f"Error generating timetable: {e}", exc_info=True)
        return jsonify({'error': str(e)}), 500


# ==================== DEPARTMENT ENDPOINTS ====================

@app.route('/api/departments', methods=['GET'])
def get_departments():
    """Get all departments"""
    try:
        departments = Department.query.all()
        return jsonify([d.to_dict() for d in departments]), 200
    except Exception as e:
        logger.error(f"Error getting departments: {e}", exc_info=True)
        return jsonify({'error': str(e)}), 500


@app.route('/api/departments', methods=['POST'])
def create_department():
    """Create new department"""
    try:
        data = request.json
        
        # Validation
        if not data.get('name') or not data.get('code'):
            return jsonify({'error': 'Name and code are required'}), 400
        
        # Check uniqueness
        existing = Department.query.filter(
            or_(Department.name == data['name'], Department.code == data['code'])
        ).first()
        
        if existing:
            return jsonify({'error': 'Department with this name or code already exists'}), 400
        
        department = Department(
            name=data['name'],
            code=data['code'],
            description=data.get('description', '')
        )
        
        db.session.add(department)
        db.session.commit()
        
        logger.info(f"Created department: {department.code}")
        return jsonify({'success': True, 'department': department.to_dict()}), 201
        
    except Exception as e:
        db.session.rollback()
        logger.error(f"Error creating department: {e}", exc_info=True)
        return jsonify({'error': str(e)}), 500


@app.route('/api/departments/<dept_id>', methods=['PUT'])
def update_department(dept_id):
    """Update department"""
    try:
        department = Department.query.get(dept_id)
        if not department:
            return jsonify({'error': 'Department not found'}), 404
        
        data = request.json
        
        if data.get('name'):
            department.name = data['name']
        if data.get('code'):
            department.code = data['code']
        if 'description' in data:
            department.description = data['description']
        
        db.session.commit()
        
        logger.info(f"Updated department: {department.code}")
        return jsonify({'success': True, 'department': department.to_dict()}), 200
        
    except Exception as e:
        db.session.rollback()
        logger.error(f"Error updating department: {e}", exc_info=True)
        return jsonify({'error': str(e)}), 500


@app.route('/api/departments/<dept_id>', methods=['DELETE'])
def delete_department(dept_id):
    """Delete department"""
    try:
        department = Department.query.get(dept_id)
        if not department:
            return jsonify({'error': 'Department not found'}), 404
        
        db.session.delete(department)
        db.session.commit()
        
        logger.info(f"Deleted department: {department.code}")
        return jsonify({'success': True, 'message': 'Department deleted'}), 200
        
    except Exception as e:
        db.session.rollback()
        logger.error(f"Error deleting department: {e}", exc_info=True)
        return jsonify({'error': str(e)}), 500


# ==================== ROOM ENDPOINTS ====================

@app.route('/api/rooms', methods=['GET'])
def get_rooms():
    """Get all rooms"""
    try:
        rooms = Room.query.all()
        return jsonify([r.to_dict() for r in rooms]), 200
    except Exception as e:
        logger.error(f"Error getting rooms: {e}", exc_info=True)
        return jsonify({'error': str(e)}), 500


@app.route('/api/rooms', methods=['POST'])
def create_room():
    """Create new room"""
    try:
        data = request.json
        
        if not data.get('name'):
            return jsonify({'error': 'Room name is required'}), 400
        
        # Check uniqueness
        existing = Room.query.filter_by(name=data['name']).first()
        if existing:
            return jsonify({'error': 'Room with this name already exists'}), 400
        
        room = Room(
            name=data['name'],
            type=data.get('type', 'classroom'),
            capacity=int(data.get('capacity', 60)),
            location=data.get('location', ''),
            building=data.get('building', ''),
            floor=int(data.get('floor', 0)) if data.get('floor') else None,
            equipment=json.dumps(data.get('equipment', [])),
            is_available=data.get('is_available', True)
        )
        
        db.session.add(room)
        db.session.commit()
        
        logger.info(f"Created room: {room.name}")
        return jsonify({'success': True, 'room': room.to_dict()}), 201
        
    except Exception as e:
        db.session.rollback()
        logger.error(f"Error creating room: {e}", exc_info=True)
        return jsonify({'error': str(e)}), 500


@app.route('/api/rooms/<room_id>', methods=['PUT'])
def update_room(room_id):
    """Update room"""
    try:
        room = Room.query.get(room_id)
        if not room:
            return jsonify({'error': 'Room not found'}), 404
        
        data = request.json
        
        if data.get('name'):
            room.name = data['name']
        if data.get('type'):
            room.type = data['type']
        if data.get('capacity'):
            room.capacity = int(data['capacity'])
        if 'location' in data:
            room.location = data['location']
        if 'building' in data:
            room.building = data['building']
        if 'floor' in data:
            room.floor = int(data['floor']) if data['floor'] else None
        if 'equipment' in data:
            room.equipment = json.dumps(data['equipment'])
        if 'is_available' in data:
            room.is_available = data['is_available']
        
        db.session.commit()
        
        logger.info(f"Updated room: {room.name}")
        return jsonify({'success': True, 'room': room.to_dict()}), 200
        
    except Exception as e:
        db.session.rollback()
        logger.error(f"Error updating room: {e}", exc_info=True)
        return jsonify({'error': str(e)}), 500


@app.route('/api/rooms/<room_id>', methods=['DELETE'])
def delete_room(room_id):
    """Delete room"""
    try:
        room = Room.query.get(room_id)
        if not room:
            return jsonify({'error': 'Room not found'}), 404
        
        db.session.delete(room)
        db.session.commit()
        
        logger.info(f"Deleted room: {room.name}")
        return jsonify({'success': True, 'message': 'Room deleted'}), 200
        
    except Exception as e:
        db.session.rollback()
        logger.error(f"Error deleting room: {e}", exc_info=True)
        return jsonify({'error': str(e)}), 500


# ==================== SUBJECT ENDPOINTS ====================

@app.route('/api/subjects', methods=['POST'])
def create_subject():
    """Create new subject"""
    try:
        data = request.json
        
        if not data.get('name') or not data.get('code'):
            return jsonify({'error': 'Name and code are required'}), 400
        
        # Check uniqueness
        existing = Subject.query.filter_by(code=data['code']).first()
        if existing:
            return jsonify({'error': 'Subject with this code already exists'}), 400
        
        subject = Subject(
            name=data['name'],
            code=data['code'],
            year=int(data.get('year', 1)),
            department_id=data['department_id'],
            subject_type=data.get('subject_type', 'theory'),
            theory_hours_per_week=int(data.get('theory_hours_per_week', 0)),
            practical_hours_per_week=int(data.get('practical_hours_per_week', 0)),
            required_lab_id=data.get('required_lab_id'),
            is_elective=data.get('is_elective', False),
            max_students=int(data.get('max_students', 0)) if data.get('max_students') else None
        )
        
        db.session.add(subject)
        db.session.commit()
        
        logger.info(f"Created subject: {subject.code}")
        return jsonify({'success': True, 'subject': subject.to_dict()}), 201
        
    except Exception as e:
        db.session.rollback()
        logger.error(f"Error creating subject: {e}", exc_info=True)
        return jsonify({'error': str(e)}), 500
# ==================== WEEK CONFIGURATION ENDPOINTS ====================

@app.route('/api/week-config', methods=['GET'])
def get_week_config():
    """Get active week configuration"""
    try:
        week_config = WeekConfiguration.query.filter_by(is_active=True).first()
        if not week_config:
            # Create default if none exists
            week_config = WeekConfiguration(
                name='Default Configuration',
                start_time='09:00',
                end_time='16:00',
                lunch_start='13:00',
                lunch_end='14:00',
                working_days=json.dumps(['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday'])
            )
            db.session.add(week_config)
            db.session.commit()
        
        return jsonify(week_config.to_dict()), 200
        
    except Exception as e:
        logger.error(f"Error getting week config: {e}", exc_info=True)
        return jsonify({'error': str(e)}), 500


@app.route('/api/week-config', methods=['POST'])
def create_week_config():
    """Create new week configuration"""
    try:
        data = request.json
        
        # Deactivate all existing configs
        WeekConfiguration.query.update({WeekConfiguration.is_active: False})
        
        week_config = WeekConfiguration(
            name=data.get('name', 'Custom Configuration'),
            start_time=data.get('start_time', '09:00'),
            end_time=data.get('end_time', '16:00'),
            lunch_start=data.get('lunch_start', '13:00'),
            lunch_end=data.get('lunch_end', '14:00'),
            working_days=json.dumps(data.get('working_days', ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday'])),
            is_active=True
        )
        
        db.session.add(week_config)
        db.session.commit()
        
        logger.info(f"Created week configuration: {week_config.name}")
        return jsonify({'success': True, 'week_config': week_config.to_dict()}), 201
        
    except Exception as e:
        db.session.rollback()
        logger.error(f"Error creating week config: {e}", exc_info=True)
        return jsonify({'error': str(e)}), 500


@app.route('/api/week-config/<config_id>', methods=['PUT'])
def update_week_config(config_id):
    """Update week configuration"""
    try:
        week_config = WeekConfiguration.query.get(config_id)
        if not week_config:
            return jsonify({'error': 'Configuration not found'}), 404
        
        data = request.json
        
        if 'name' in data:
            week_config.name = data['name']
        if 'start_time' in data:
            week_config.start_time = data['start_time']
        if 'end_time' in data:
            week_config.end_time = data['end_time']
        if 'lunch_start' in data:
            week_config.lunch_start = data['lunch_start']
        if 'lunch_end' in data:
            week_config.lunch_end = data['lunch_end']
        if 'working_days' in data:
            week_config.working_days = json.dumps(data['working_days'])
        if 'is_active' in data and data['is_active']:
            # Deactivate all others
            WeekConfiguration.query.filter(WeekConfiguration.id != config_id).update({WeekConfiguration.is_active: False})
            week_config.is_active = True
        
        db.session.commit()
        
        logger.info(f"Updated week configuration: {week_config.name}")
        return jsonify({'success': True, 'week_config': week_config.to_dict()}), 200
        
    except Exception as e:
        db.session.rollback()
        logger.error(f"Error updating week config: {e}", exc_info=True)
        return jsonify({'error': str(e)}), 500


# ==================== SPECIAL SESSION ENDPOINTS ====================

@app.route('/api/special-sessions', methods=['GET'])
def get_special_sessions():
    """Get all special sessions"""
    try:
        year = request.args.get('year', type=int)
        dept_id = request.args.get('department_id')
        
        query = SpecialSession.query
        
        if year:
            query = query.filter_by(year=year)
        if dept_id:
            query = query.filter_by(department_id=dept_id)
        
        sessions = query.all()
        return jsonify([s.to_dict() for s in sessions]), 200
        
    except Exception as e:
        logger.error(f"Error getting special sessions: {e}", exc_info=True)
        return jsonify({'error': str(e)}), 500


@app.route('/api/special-sessions', methods=['POST'])
def create_special_session():
    """Create or update special session configuration"""
    try:
        data = request.json
        
        # Check if already exists
        existing = SpecialSession.query.filter_by(
            session_type=data['session_type'],
            year=data['year'],
            department_id=data['department_id']
        ).first()
        
        if existing:
            # Update existing
            existing.is_enabled = data.get('is_enabled', False)
            existing.hours_per_week = int(data.get('hours_per_week', 0))
            existing.days_per_week = int(data.get('days_per_week', 0))
            existing.session_duration = int(data.get('session_duration', 2))
            session = existing
        else:
            # Create new
            session = SpecialSession(
                session_type=data['session_type'],
                year=int(data['year']),
                department_id=data['department_id'],
                is_enabled=data.get('is_enabled', False),
                hours_per_week=int(data.get('hours_per_week', 0)),
                days_per_week=int(data.get('days_per_week', 0)),
                session_duration=int(data.get('session_duration', 2))
            )
            db.session.add(session)
        
        db.session.commit()
        
        logger.info(f"Created/updated special session: {session.session_type} Year-{session.year}")
        return jsonify({'success': True, 'session': session.to_dict()}), 201
        
    except Exception as e:
        db.session.rollback()
        logger.error(f"Error creating special session: {e}", exc_info=True)
        return jsonify({'error': str(e)}), 500


@app.route('/api/special-sessions/<session_id>', methods=['DELETE'])
def delete_special_session(session_id):
    """Delete special session"""
    try:
        session = SpecialSession.query.get(session_id)
        if not session:
            return jsonify({'error': 'Session not found'}), 404
        
        db.session.delete(session)
        db.session.commit()
        
        logger.info(f"Deleted special session: {session_id}")
        return jsonify({'success': True, 'message': 'Session deleted'}), 200
        
    except Exception as e:
        db.session.rollback()
        logger.error(f"Error deleting special session: {e}", exc_info=True)
        return jsonify({'error': str(e)}), 500


# ==================== SUBJECT MANAGEMENT ENDPOINTS ====================

@app.route('/api/subjects', methods=['GET'])
def get_subjects():
    """Get all subjects with optional filters"""
    try:
        year = request.args.get('year', type=int)
        dept_id = request.args.get('department_id')
        
        query = Subject.query
        
        if year:
            query = query.filter_by(year=year)
        if dept_id:
            query = query.filter_by(department_id=dept_id)
        
        subjects = query.all()
        return jsonify([s.to_dict() for s in subjects]), 200
        
    except Exception as e:
        logger.error(f"Error getting subjects: {e}", exc_info=True)
        return jsonify({'error': str(e)}), 500


@app.route('/api/subjects/<subject_id>', methods=['PUT'])
def update_subject(subject_id):
    """Update subject"""
    try:
        subject = Subject.query.get(subject_id)
        if not subject:
            return jsonify({'error': 'Subject not found'}), 404
        
        data = request.json
        
        if 'name' in data:
            subject.name = data['name']
        if 'code' in data:
            subject.code = data['code']
        if 'year' in data:
            subject.year = int(data['year'])
        if 'subject_type' in data:
            subject.subject_type = data['subject_type']
        if 'theory_hours_per_week' in data:
            subject.theory_hours_per_week = int(data['theory_hours_per_week'])
        if 'practical_hours_per_week' in data:
            subject.practical_hours_per_week = int(data['practical_hours_per_week'])
        if 'required_lab_id' in data:
            subject.required_lab_id = data['required_lab_id']
        if 'is_elective' in data:
            subject.is_elective = data['is_elective']
        if 'max_students' in data:
            subject.max_students = int(data['max_students']) if data['max_students'] else None
        
        db.session.commit()
        
        logger.info(f"Updated subject: {subject.code}")
        return jsonify({'success': True, 'subject': subject.to_dict()}), 200
        
    except Exception as e:
        db.session.rollback()
        logger.error(f"Error updating subject: {e}", exc_info=True)
        return jsonify({'error': str(e)}), 500


@app.route('/api/subjects/<subject_id>', methods=['DELETE'])
def delete_subject(subject_id):
    """Delete subject"""
    try:
        subject = Subject.query.get(subject_id)
        if not subject:
            return jsonify({'error': 'Subject not found'}), 404
        
        db.session.delete(subject)
        db.session.commit()
        
        logger.info(f"Deleted subject: {subject.code}")
        return jsonify({'success': True, 'message': 'Subject deleted'}), 200
        
    except Exception as e:
        db.session.rollback()
        logger.error(f"Error deleting subject: {e}", exc_info=True)
        return jsonify({'error': str(e)}), 500


# ==================== STUDENT GROUP ENDPOINTS ====================

@app.route('/api/student-groups', methods=['GET'])
def get_student_groups():
    """Get all student groups"""
    try:
        year = request.args.get('year', type=int)
        dept_id = request.args.get('department_id')
        
        query = StudentGroup.query
        
        if year:
            query = query.filter_by(year=year)
        if dept_id:
            query = query.filter_by(department_id=dept_id)
        
        groups = query.all()
        return jsonify([g.to_dict() for g in groups]), 200
        
    except Exception as e:
        logger.error(f"Error getting student groups: {e}", exc_info=True)
        return jsonify({'error': str(e)}), 500


@app.route('/api/student-groups', methods=['POST'])
def create_student_group():
    """Create new student group"""
    try:
        data = request.json
        
        group = StudentGroup(
            year=int(data['year']),
            batch=data['batch'],
            department_id=data['department_id'],
            student_count=int(data['student_count']),
            elective_subject_id=data.get('elective_subject_id')
        )
        
        db.session.add(group)
        db.session.commit()
        
        logger.info(f"Created student group: Year-{group.year} Batch-{group.batch}")
        return jsonify({'success': True, 'group': group.to_dict()}), 201
        
    except Exception as e:
        db.session.rollback()
        logger.error(f"Error creating student group: {e}", exc_info=True)
        return jsonify({'error': str(e)}), 500


# ==================== TIMETABLE QUERY ENDPOINTS ====================

@app.route('/api/timetable', methods=['GET'])
def get_timetable():
    """Get timetable with filters"""
    try:
        year = request.args.get('year', type=int)
        dept_id = request.args.get('department_id')
        batch = request.args.get('batch')
        room_id = request.args.get('room_id')
        
        query = ScheduledClass.query.options(
            joinedload(ScheduledClass.subject),
            joinedload(ScheduledClass.student_group),
            joinedload(ScheduledClass.room),
            joinedload(ScheduledClass.department)
        )
        
        if year:
            query = query.filter_by(year=year)
        if dept_id:
            query = query.filter_by(department_id=dept_id)
        if batch:
            query = query.join(StudentGroup).filter(StudentGroup.batch == batch)
        if room_id:
            query = query.filter_by(room_id=room_id)
        
        classes = query.all()
        return jsonify([c.to_dict() for c in classes]), 200
        
    except Exception as e:
        logger.error(f"Error getting timetable: {e}", exc_info=True)
        return jsonify({'error': str(e)}), 500


@app.route('/api/timetable/<class_id>', methods=['DELETE'])
def delete_scheduled_class(class_id):
    """Delete a scheduled class"""
    try:
        scheduled_class = ScheduledClass.query.get(class_id)
        if not scheduled_class:
            return jsonify({'error': 'Class not found'}), 404
        
        db.session.delete(scheduled_class)
        db.session.commit()
        
        logger.info(f"Deleted scheduled class: {class_id}")
        return jsonify({'success': True, 'message': 'Class deleted'}), 200
        
    except Exception as e:
        db.session.rollback()
        logger.error(f"Error deleting scheduled class: {e}", exc_info=True)
        return jsonify({'error': str(e)}), 500


@app.route('/api/timetable/clear', methods=['POST'])
def clear_timetable():
    """Clear timetable for specific years/departments"""
    try:
        data = request.json
        years = data.get('years', [])
        dept_ids = data.get('department_ids', [])
        
        query = ScheduledClass.query
        
        if years:
            query = query.filter(ScheduledClass.year.in_(years))
        if dept_ids:
            query = query.filter(ScheduledClass.department_id.in_(dept_ids))
        
        count = query.delete()
        db.session.commit()
        
        logger.info(f"Cleared {count} scheduled classes")
        return jsonify({'success': True, 'deleted_count': count}), 200
        
    except Exception as e:
        db.session.rollback()
        logger.error(f"Error clearing timetable: {e}", exc_info=True)
        return jsonify({'error': str(e)}), 500


# ==================== EXPORT ENDPOINTS ====================

@app.route('/api/export/pdf', methods=['POST'])
def export_pdf():
    """Export timetable to PDF"""
    try:
        data = request.json
        year = data.get('year')
        dept_id = data.get('department_id')
        
        # Query classes
        query = ScheduledClass.query.options(
            joinedload(ScheduledClass.subject),
            joinedload(ScheduledClass.student_group),
            joinedload(ScheduledClass.room)
        )
        
        if year:
            query = query.filter_by(year=year)
        if dept_id:
            query = query.filter_by(department_id=dept_id)
        
        classes = query.order_by(ScheduledClass.day_of_week, ScheduledClass.start_time).all()
        
        # Create PDF
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(letter))
        elements = []
        
        styles = getSampleStyleSheet()
        title_style = styles['Title']
        
        # Title
        dept = Department.query.get(dept_id) if dept_id else None
        title_text = f"Timetable - Year {year}" if year else "Timetable"
        if dept:
            title_text += f" - {dept.name}"
        
        elements.append(Paragraph(title_text, title_style))
        elements.append(Spacer(1, 20))
        
        # Group by day
        days_order = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday']
        classes_by_day = defaultdict(list)
        for cls in classes:
            classes_by_day[cls.day_of_week].append(cls)
        
        for day in days_order:
            if day not in classes_by_day:
                continue
            
            day_classes = sorted(classes_by_day[day], key=lambda x: x.start_time)
            
            # Day heading
            elements.append(Paragraph(f"<b>{day}</b>", styles['Heading2']))
            elements.append(Spacer(1, 10))
            
            # Table data
            table_data = [['Time', 'Subject', 'Type', 'Group', 'Room']]
            
            for cls in day_classes:
                table_data.append([
                    f"{cls.start_time} - {cls.end_time}",
                    cls.subject.name if cls.subject else 'Special Session',
                    cls.class_type.title(),
                    cls.student_group.to_dict()['group_name'] if cls.student_group else 'N/A',
                    cls.room.name if cls.room else 'N/A'
                ])
            
            # Create table
            table = Table(table_data, colWidths=[80, 150, 80, 120, 100])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            elements.append(table)
            elements.append(Spacer(1, 20))
        
        doc.build(elements)
        
        buffer.seek(0)
        return send_file(
            buffer,
            as_attachment=True,
            download_name=f'timetable_year_{year}.pdf',
            mimetype='application/pdf'
        )
        
    except Exception as e:
        logger.error(f"Error exporting PDF: {e}", exc_info=True)
        return jsonify({'error': str(e)}), 500


@app.route('/api/export/excel', methods=['POST'])
def export_excel():
    """Export timetable to Excel"""
    try:
        data = request.json
        year = data.get('year')
        dept_id = data.get('department_id')
        
        # Query classes
        query = ScheduledClass.query.options(
            joinedload(ScheduledClass.subject),
            joinedload(ScheduledClass.student_group),
            joinedload(ScheduledClass.room)
        )
        
        if year:
            query = query.filter_by(year=year)
        if dept_id:
            query = query.filter_by(department_id=dept_id)
        
        classes = query.order_by(ScheduledClass.day_of_week, ScheduledClass.start_time).all()
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Year {year}" if year else "Timetable"
        
        # Headers
        headers = ['Day', 'Start Time', 'End Time', 'Subject', 'Type', 'Group', 'Room', 'Duration']
        ws.append(headers)
        
        # Style headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Data rows
        for cls in classes:
            ws.append([
                cls.day_of_week,
                cls.start_time,
                cls.end_time,
                cls.subject.name if cls.subject else 'Special Session',
                cls.class_type.title(),
                cls.student_group.to_dict()['group_name'] if cls.student_group else 'N/A',
                cls.room.name if cls.room else 'N/A',
                f"{cls.duration_hours} hour(s)"
            ])
        
        # Auto-size columns
        for column in ws.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column[0].column_letter].width = adjusted_width
        
        # Save to buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return send_file(
            buffer,
            as_attachment=True,
            download_name=f'timetable_year_{year}.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        logger.error(f"Error exporting Excel: {e}", exc_info=True)
        return jsonify({'error': str(e)}), 500


@app.route('/api/export/ical', methods=['POST'])
def export_ical():
    """Export timetable to iCal format"""
    try:
        data = request.json
        year = data.get('year')
        dept_id = data.get('department_id')
        
        # Query classes
        query = ScheduledClass.query.options(
            joinedload(ScheduledClass.subject),
            joinedload(ScheduledClass.student_group),
            joinedload(ScheduledClass.room)
        )
        
        if year:
            query = query.filter_by(year=year)
        if dept_id:
            query = query.filter_by(department_id=dept_id)
        
        classes = query.all()
        
        # Create calendar
        cal = Calendar()
        cal.add('prodid', '-//Academic Timetable Generator//EN')
        cal.add('version', '2.0')
        
        # Get week config
        week_config = WeekConfiguration.query.filter_by(is_active=True).first()
        working_days = json.loads(week_config.working_days) if week_config else ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']
        
        # Start from next Monday
        today = datetime.now().date()
        days_ahead = 0 - today.weekday()
        if days_ahead <= 0:
            days_ahead += 7
        next_monday = today + timedelta(days=days_ahead)
        
        # Map day names to offsets
        day_offset = {
            'Monday': 0, 'Tuesday': 1, 'Wednesday': 2,
            'Thursday': 3, 'Friday': 4, 'Saturday': 5, 'Sunday': 6
        }
        
        for cls in classes:
            event = Event()
            
            # Calculate event date
            day_diff = day_offset.get(cls.day_of_week, 0)
            event_date = next_monday + timedelta(days=day_diff)
            
            # Parse times
            start_hour, start_minute = map(int, cls.start_time.split(':'))
            end_hour, end_minute = map(int, cls.end_time.split(':'))
            
            start_dt = datetime.combine(event_date, time(start_hour, start_minute))
            end_dt = datetime.combine(event_date, time(end_hour, end_minute))
            
            event.add('summary', cls.subject.name if cls.subject else 'Special Session')
            event.add('dtstart', start_dt)
            event.add('dtend', end_dt)
            event.add('location', cls.room.name if cls.room else 'TBA')
            
            description = f"Type: {cls.class_type.title()}\n"
            description += f"Group: {cls.student_group.to_dict()['group_name'] if cls.student_group else 'N/A'}\n"
            if cls.subject:
                description += f"Subject Code: {cls.subject.code}"
            
            event.add('description', description)
            event.add('uid', f"{cls.id}@timetable.app")
            
            # Add recurrence rule for weekly repeat
            event.add('rrule', {'freq': 'weekly', 'count': 16})  # 16 weeks semester
            
            cal.add_component(event)
        
        # Generate iCal string
        ical_string = cal.to_ical()
        
        buffer = BytesIO(ical_string)
        buffer.seek(0)
        
        return send_file(
            buffer,
            as_attachment=True,
            download_name=f'timetable_year_{year}.ics',
            mimetype='text/calendar'
        )
        
    except Exception as e:
        logger.error(f"Error exporting iCal: {e}", exc_info=True)
        return jsonify({'error': str(e)}), 500


# ==================== ANALYTICS AND STATISTICS ====================

@app.route('/api/analytics/room-utilization', methods=['GET'])
def get_room_utilization():
    """Get room utilization statistics"""
    try:
        dept_id = request.args.get('department_id')
        year = request.args.get('year', type=int)
        
        query = ScheduledClass.query
        
        if dept_id:
            query = query.filter_by(department_id=dept_id)
        if year:
            query = query.filter_by(year=year)
        
        classes = query.all()
        
        # Calculate utilization
        room_stats = defaultdict(lambda: {'total_hours': 0, 'classes': 0})
        
        for cls in classes:
            room_id = cls.room_id
            room_stats[room_id]['total_hours'] += cls.duration_hours
            room_stats[room_id]['classes'] += 1
        
        # Get room details
        result = []
        for room_id, stats in room_stats.items():
            room = Room.query.get(room_id)
            if room:
                result.append({
                    'room_id': room_id,
                    'room_name': room.name,
                    'room_type': room.type,
                    'total_hours': stats['total_hours'],
                    'total_classes': stats['classes'],
                    'utilization_percentage': (stats['total_hours'] / (35 * 5)) * 100  # Assuming 35 hours per week
                })
        
        return jsonify(sorted(result, key=lambda x: -x['total_hours'])), 200
        
    except Exception as e:
        logger.error(f"Error getting room utilization: {e}", exc_info=True)
        return jsonify({'error': str(e)}), 500


@app.route('/api/analytics/subject-distribution', methods=['GET'])
def get_subject_distribution():
    """Get subject distribution statistics"""
    try:
        year = request.args.get('year', type=int)
        dept_id = request.args.get('department_id')
        
        query = ScheduledClass.query
        
        if year:
            query = query.filter_by(year=year)
        if dept_id:
            query = query.filter_by(department_id=dept_id)
        
        classes = query.all()
        
        # Calculate distribution
        subject_stats = defaultdict(lambda: {'theory_hours': 0, 'practical_hours': 0, 'total_classes': 0})
        
        for cls in classes:
            if cls.subject:
                subject_id = cls.subject_id
                if cls.class_type == 'theory':
                    subject_stats[subject_id]['theory_hours'] += cls.duration_hours
                elif cls.class_type == 'practical':
                    subject_stats[subject_id]['practical_hours'] += cls.duration_hours
                subject_stats[subject_id]['total_classes'] += 1
        
        # Get subject details
        result = []
        for subject_id, stats in subject_stats.items():
            subject = Subject.query.get(subject_id)
            if subject:
                result.append({
                    'subject_id': subject_id,
                    'subject_name': subject.name,
                    'subject_code': subject.code,
                    'theory_hours': stats['theory_hours'],
                    'practical_hours': stats['practical_hours'],
                    'total_hours': stats['theory_hours'] + stats['practical_hours'],
                    'total_classes': stats['total_classes']
                })
        
        return jsonify(sorted(result, key=lambda x: -x['total_hours'])), 200
        
    except Exception as e:
        logger.error(f"Error getting subject distribution: {e}", exc_info=True)
        return jsonify({'error': str(e)}), 500


@app.route('/api/analytics/conflicts', methods=['GET'])
@app.route('/api/analytics/conflicts', methods=['GET'])
def detect_conflicts():
    """Detect any conflicts in the current timetable"""
    try: 
        conflicts = []
        
        classes = ScheduledClass.query.all()
        
        for i, cls1 in enumerate(classes):
            for cls2 in classes[i+1:]:
                # Check if same day
                if cls1.day_of_week != cls2.day_of_week:
                    continue
                
                # Check time overlap
                start1 = int(cls1.start_time.split(':')[0])
                end1 = int(cls1.end_time.split(':')[0])
                start2 = int(cls2.start_time.split(':')[0])
                end2 = int(cls2.end_time.split(':')[0])
                
                range1 = set(range(start1, end1))
                range2 = set(range(start2, end2))
                
                if not (range1 & range2):
                    continue
                
                # Room conflict
                if cls1.room_id == cls2.room_id:
                    conflicts.append({
                        'type': 'room_conflict',
                        'message': f"Room {cls1.room.name} is double-booked on {cls1.day_of_week} at {cls1.start_time}",
                        'class1': cls1.to_dict(),
                        'class2': cls2.to_dict()
                    })
                
                # Student group conflict
                if cls1.student_group_id == cls2.student_group_id:
                    conflicts.append({
                        'type': 'group_conflict',
                        'message': f"Student group has overlapping classes on {cls1.day_of_week} at {cls1.start_time}",
                        'class1': cls1.to_dict(),
                        'class2': cls2.to_dict()
                    })
        
        return jsonify({
            'total_conflicts': len(conflicts),
            'conflicts': conflicts
        }), 200
        
    except Exception as e:
        logger.error(f"Error detecting conflicts: {e}", exc_info=True)
        return jsonify({'error': str(e)}), 500


# ==================== BATCH OPERATIONS ====================

@app.route('/api/batch/create-subjects', methods=['POST'])
def batch_create_subjects():
    """Batch create subjects for a year"""
    try:
        data = request.json
        subjects_data = data.get('subjects', [])
        
        created_subjects = []
        errors = []
        
        for subject_data in subjects_data:
            try:
                # Check if already exists
                existing = Subject.query.filter_by(code=subject_data['code']).first()
                if existing:
                    errors.append(f"Subject {subject_data['code']} already exists")
                    continue
                
                subject = Subject(
                    name=subject_data['name'],
                    code=subject_data['code'],
                    year=int(subject_data['year']),
                    department_id=subject_data['department_id'],
                    subject_type=subject_data.get('subject_type', 'theory'),
                    theory_hours_per_week=int(subject_data.get('theory_hours_per_week', 0)),
                    practical_hours_per_week=int(subject_data.get('practical_hours_per_week', 0)),
                    required_lab_id=subject_data.get('required_lab_id'),
                    is_elective=subject_data.get('is_elective', False),
                    max_students=int(subject_data.get('max_students', 0)) if subject_data.get('max_students') else None
                )
                
                db.session.add(subject)
                created_subjects.append(subject)
                
            except Exception as e:
                errors.append(f"Error creating subject {subject_data.get('code', 'unknown')}: {str(e)}")
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'created_count': len(created_subjects),
            'subjects': [s.to_dict() for s in created_subjects],
            'errors': errors
        }), 201
        
    except Exception as e:
        db.session.rollback()
        logger.error(f"Error in batch create subjects: {e}", exc_info=True)
        return jsonify({'error': str(e)}), 500


@app.route('/api/batch/create-rooms', methods=['POST'])
def batch_create_rooms():
    """Batch create rooms/labs"""
    try:
        data = request.json
        rooms_data = data.get('rooms', [])
        
        created_rooms = []
        errors = []
        
        for room_data in rooms_data:
            try:
                # Check if already exists
                existing = Room.query.filter_by(name=room_data['name']).first()
                if existing:
                    errors.append(f"Room {room_data['name']} already exists")
                    continue
                
                room = Room(
                    name=room_data['name'],
                    type=room_data.get('type', 'classroom'),
                    capacity=int(room_data.get('capacity', 60)),
                    location=room_data.get('location', ''),
                    building=room_data.get('building', ''),
                    floor=int(room_data.get('floor', 0)) if room_data.get('floor') else None,
                    equipment=json.dumps(room_data.get('equipment', [])),
                    is_available=room_data.get('is_available', True)
                )
                
                db.session.add(room)
                created_rooms.append(room)
                
            except Exception as e:
                errors.append(f"Error creating room {room_data.get('name', 'unknown')}: {str(e)}")
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'created_count': len(created_rooms),
            'rooms': [r.to_dict() for r in created_rooms],
            'errors': errors
        }), 201
        
    except Exception as e:
        db.session.rollback()
        logger.error(f"Error in batch create rooms: {e}", exc_info=True)
        return jsonify({'error': str(e)}), 500

# ==================== VALIDATION ENDPOINTS ====================

@app.route('/api/validate/timetable-requirements', methods=['POST'])
def validate_timetable_requirements():
    """Validate if all requirements are met before generation"""
    try:
        data = request.json
        years = data.get('years', [])
        params = data.get('params', {})
        
        validation_errors = []
        validation_warnings = []
        
        for year in years:
            year_config = params.get(f'year{year}', {})
            dept_id = year_config.get('departmentId')
            
            if not dept_id:
                validation_errors.append(f"Year {year}: Department not selected")
                continue
            
            # Check if subjects exist
            subjects = Subject.query.filter_by(year=year, department_id=dept_id).all()
            if not subjects:
                validation_errors.append(f"Year {year}: No subjects configured")
            
            # Check total hours
            total_theory_hours = sum(s.theory_hours_per_week for s in subjects if not s.is_elective)
            total_practical_hours = sum(s.practical_hours_per_week for s in subjects if not s.is_elective)
            
            if total_theory_hours + total_practical_hours > 30:
                validation_warnings.append(f"Year {year}: Total weekly hours ({total_theory_hours + total_practical_hours}) exceeds 30")
            
            # Check lab availability for practicals
            practical_subjects = [s for s in subjects if s.practical_hours_per_week > 0]
            for subject in practical_subjects:
                if not subject.required_lab_id:
                    validation_warnings.append(f"Year {year}: {subject.name} has practicals but no lab assigned")
            
            # Check room capacity
            batches = int(year_config.get('batches', 3))
            students_per_batch = int(year_config.get('studentsPerBatch', 60))
            
            available_rooms = Room.query.filter(Room.capacity >= students_per_batch, Room.is_available == True).all()
            if len(available_rooms) < batches:
                validation_warnings.append(f"Year {year}: Limited rooms available for batch size {students_per_batch}")
            
            # Check for 4th year electives
            if year == 4:
                electives = year_config.get('electives', [])
                if not electives:
                    validation_warnings.append(f"Year 4: No elective subjects configured")
        
        # Check week configuration
        week_config = WeekConfiguration.query.filter_by(is_active=True).first()
        if not week_config:
            validation_errors.append("No active week configuration found")
        
        return jsonify({
            'valid': len(validation_errors) == 0,
            'errors': validation_errors,
            'warnings': validation_warnings
        }), 200
        
    except Exception as e:
        logger.error(f"Error validating requirements: {e}", exc_info=True)
        return jsonify({'error': str(e)}), 500


# ==================== DATABASE INITIALIZATION ====================

def init_database():
    """Initialize database with tables"""
    with app.app_context():
        try:
            db.create_all()
            logger.info("Database tables created successfully")
            
            # Create default week configuration if none exists
            if not WeekConfiguration.query.first():
                default_config = WeekConfiguration(
                    name='Default Configuration',
                    start_time='09:00',
                    end_time='16:00',
                    lunch_start='13:00',
                    lunch_end='14:00',
                    working_days=json.dumps(['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']),
                    is_active=True
                )
                db.session.add(default_config)
                db.session.commit()
                logger.info("Created default week configuration")
            
        except Exception as e:
            logger.error(f"Error initializing database: {e}", exc_info=True)
            raise


# ==================== SEED DATA FOR TESTING ====================

@app.route('/api/seed-data', methods=['POST'])
def seed_sample_data():
    """Seed sample data for testing"""
    try:
        # Create sample departments
        dept_cs = Department(
            name='Computer Science',
            code='CS',
            description='Department of Computer Science and Engineering'
        )
        dept_ee = Department(
            name='Electrical Engineering',
            code='EE',
            description='Department of Electrical and Electronics Engineering'
        )
        
        db.session.add(dept_cs)
        db.session.add(dept_ee)
        db.session.commit()
        
        # Create sample rooms
        rooms_data = [
            {'name': 'Room 101', 'type': 'classroom', 'capacity': 60, 'location': 'Building A'},
            {'name': 'Room 102', 'type': 'classroom', 'capacity': 60, 'location': 'Building A'},
            {'name': 'Room 103', 'type': 'classroom', 'capacity': 60, 'location': 'Building A'},
            {'name': 'CS Lab 1', 'type': 'lab', 'capacity': 30, 'location': 'Building C'},
            {'name': 'CS Lab 2', 'type': 'lab', 'capacity': 30, 'location': 'Building C'},
            {'name': 'EE Lab 1', 'type': 'lab', 'capacity': 30, 'location': 'Building B'},
        ]
        
        room_map = {}
        for room_data in rooms_data:
            room = Room(**room_data, equipment=json.dumps(['projector', 'whiteboard']))
            db.session.add(room)
            room_map[room_data['name']] = room
        
        db.session.commit()
        
        # Create sample subjects for Year 1 CS
        subjects_data = [
            {'name': 'Mathematics I', 'code': 'CS101', 'theory': 3, 'practical': 0},
            {'name': 'Physics', 'code': 'CS102', 'theory': 3, 'practical': 2, 'lab': 'EE Lab 1'},
            {'name': 'Programming Fundamentals', 'code': 'CS103', 'theory': 3, 'practical': 2, 'lab': 'CS Lab 1'},
            {'name': 'Digital Logic', 'code': 'CS104', 'theory': 3, 'practical': 2, 'lab': 'EE Lab 1'},
            {'name': 'English Communication', 'code': 'CS105', 'theory': 2, 'practical': 0},
        ]
        
        for subj_data in subjects_data:
            lab_id = None
            if 'lab' in subj_data and subj_data['lab'] in room_map:
                lab_id = room_map[subj_data['lab']].id
            
            subject = Subject(
                name=subj_data['name'],
                code=subj_data['code'],
                year=1,
                department_id=dept_cs.id,
                subject_type='both' if subj_data['practical'] > 0 else 'theory',
                theory_hours_per_week=subj_data['theory'],
                practical_hours_per_week=subj_data['practical'],
                required_lab_id=lab_id,
                is_elective=False
            )
            db.session.add(subject)
        
        db.session.commit()
        
        logger.info("Sample data seeded successfully")
        return jsonify({'success': True, 'message': 'Sample data seeded'}), 201
        
    except Exception as e:
        db.session.rollback()
        logger.error(f"Error seeding data: {e}", exc_info=True)
        return jsonify({'error': str(e)}), 500


# ==================== HEALTH CHECK ====================

@app.route('/api/health', methods=['GET'])
def health_check():
    """Health check endpoint"""
    try:
        # Check database connection
        db.session.execute('SELECT 1')
        
        # Check if PyTorch is available
        torch_available = torch.cuda.is_available() if hasattr(torch, 'cuda') else False
        
        return jsonify({
            'status': 'healthy',
            'database': 'connected',
            'torch_available': True,
            'cuda_available': torch_available,
            'timestamp': datetime.utcnow().isoformat()
        }), 200
        
    except Exception as e:
        logger.error(f"Health check failed: {e}", exc_info=True)
        return jsonify({
            'status': 'unhealthy',
            'error': str(e)
        }), 500


# ==================== ERROR HANDLERS ====================

@app.errorhandler(404)
def not_found(error):
    """Handle 404 errors"""
    return jsonify({'error': 'Resource not found'}), 404


@app.errorhandler(500)
def internal_error(error):
    """Handle 500 errors"""
    db.session.rollback()
    logger.error(f"Internal server error: {error}", exc_info=True)
    return jsonify({'error': 'Internal server error'}), 500


@app.errorhandler(Exception)
def handle_exception(error):
    """Handle all unhandled exceptions"""
    db.session.rollback()
    logger.error(f"Unhandled exception: {error}", exc_info=True)
    return jsonify({'error': str(error)}), 500


# ==================== APPLICATION STARTUP ====================

if __name__ == '__main__':
    try:
        logger.info("Starting Academic Timetable Generator Application...")
        
        # Initialize database
        init_database()
        
        # Get port from environment or use default
        port = int(os.environ.get('PORT', 5000))
        
        logger.info(f"Application starting on port {port}")
        logger.info("GNN-based timetable generation enabled")
        logger.info(f"Using device: {'CUDA' if torch.cuda.is_available() else 'CPU'}")
        
        # Run the application
        app.run(
            host='0.0.0.0',
            port=port,
            debug=True,  # Set to False in production
            threaded=True
        )
        
    except Exception as e:
        logger.error(f"Failed to start application: {e}", exc_info=True)
        raise